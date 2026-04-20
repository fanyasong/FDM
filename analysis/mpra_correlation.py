"""
Reproduce MPRA correlation results (Table 5).

Correlates FDM L8 cross-layer-normalised surprisal with lentiMPRA
enhancer activity from Agarwal et al. Nature 2025 (Zenodo: 10558183).

Usage:
    python analysis/mpra_correlation.py \
        --checkpoint checkpoints/entropy_pretrain/best.pt \
        --seq_table data/mpra/table_s3.xlsx \
        --act_table data/mpra/table_s6.xlsx \
        --cell_type HepG2

Expected output (Table 5):
    HepG2: raw r=0.176  GC-stratified r=0.096  n=61,463
    K562:  raw r=0.124  GC-stratified r=0.110  n=2,123
    Random model K562: GC-stratified r=-0.049  [null control]
"""

import argparse
import sys
import numpy as np
import openpyxl
import torch
from scipy.stats import spearmanr
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from fdm import FDM_Entropy_LM

DNA_VOCAB = {b: i for i, b in enumerate("ATGCN")}
DNA_VOCAB.update({b.lower(): i for i, b in enumerate("ATGCN")})
SEQ_LEN   = 512   # FDM input length; MPRA sequences are 200bp, padded with N


def load_mpra_sequences(seq_table: str, sheet_name: str) -> dict:
    """Load sequences from Table S3 of Agarwal et al. 2025."""
    wb = openpyxl.load_workbook(seq_table, read_only=True)
    seqs = {}
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] is None:
            continue
        if str(row[0]) in ("name", "Elements were designed"):
            continue
        if "_Reversed" in str(row[0]):
            continue
        name = str(row[0])
        seq = None
        for val in row:
            if isinstance(val, str) and len(val) >= 200:
                if set(val.upper()).issubset({"A", "T", "G", "C", "N"}):
                    # Strip 15nt adaptor on each side if 230bp
                    seq = val[15:215] if len(val) >= 230 else val[:200]
                    break
        if seq:
            seqs[name] = seq
    wb.close()
    return seqs


def load_mpra_activity(act_table: str, cell_type: str) -> dict:
    """Load log2(RNA/DNA) activity from Table S6."""
    wb = openpyxl.load_workbook(act_table, read_only=True)
    sheet = f"Fold assignment, {cell_type}"
    acts = {}
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[1] and row[2] is not None:
            acts[str(row[1])] = float(row[2])
    wb.close()
    return acts


def encode(seq: str, pad_to: int = SEQ_LEN) -> list:
    ids = [DNA_VOCAB.get(b, 4) for b in seq.upper()]
    ids += [4] * (pad_to - len(ids))
    return ids[:pad_to]


@torch.no_grad()
def get_p_scores(model, seq_dict: dict, ids: list, device: str, batch_size=128) -> np.ndarray:
    """Compute per-sequence L8 cross-layer-normalised surprisal."""
    out = []
    model = model.to(device)
    model.eval()
    for i in range(0, len(ids), batch_size):
        batch = ids[i:i+batch_size]
        xs = [encode(seq_dict[sid]) for sid in batch]
        x = torch.tensor(xs, dtype=torch.long, device=device)
        sc = model.get_surprisal_scores(x, normalize_layers=True)
        # Mean over all positions → scalar per sequence
        out.extend(sc.mean(dim=1)[:, 7].cpu().tolist())
    return np.array(out)


def gc_stratified_r(p: np.ndarray, activity: np.ndarray, gc: np.ndarray, n_bins=5) -> float:
    """Compute GC-stratified Spearman r (averaged within GC quintiles)."""
    quantiles = np.linspace(0, 100, n_bins + 1)
    bin_edges = np.percentile(gc, quantiles)
    rs = []
    for j in range(n_bins):
        mask = (gc >= bin_edges[j]) & (gc < bin_edges[j+1])
        if mask.sum() < 50:
            continue
        r, _ = spearmanr(p[mask], activity[mask])
        rs.append(r)
    return float(np.mean(rs)) if rs else float("nan")


def main(args):
    print(f"Loading model from {args.checkpoint} ...", flush=True)
    model = FDM_Entropy_LM()
    ckpt = torch.load(args.checkpoint, map_location="cpu", weights_only=False)
    model.load_state_dict(ckpt["model_state"], strict=False)

    model_rand = FDM_Entropy_LM()

    cell_types = [args.cell_type] if args.cell_type != "all" else ["HepG2", "K562"]
    sheet_map = {"HepG2": "HepG2 large-scale", "K562": "K562 large-scale"}

    for cell in cell_types:
        print(f"\n{'='*50}")
        print(f"Cell type: {cell}", flush=True)

        seqs  = load_mpra_sequences(args.seq_table, sheet_map[cell])
        acts  = load_mpra_activity(args.act_table, cell)
        common = [k for k in seqs if k in acts]
        print(f"  Sequences with activity: {len(common)}", flush=True)

        activity = np.array([acts[k]  for k in common])
        gc       = np.array([(seqs[k].upper().count("G") + seqs[k].upper().count("C")) / 200.0
                              for k in common])

        print(f"  GC range: {gc.min():.2f}–{gc.max():.2f}  mean={gc.mean():.2f}", flush=True)

        print(f"  Computing FDM scores ...", flush=True)
        p_trained = get_p_scores(model, seqs, common, args.device)

        r_raw, _ = spearmanr(p_trained, activity)
        r_gc     = gc_stratified_r(p_trained, activity, gc)

        # GC-corrected residual
        gc_coef = np.polyfit(gc, activity, 1)
        act_resid = activity - (gc_coef[0]*gc + gc_coef[1])
        gc_coef2  = np.polyfit(gc, p_trained, 1)
        p_resid   = p_trained - (gc_coef2[0]*gc + gc_coef2[1])
        r_corr, p_corr = spearmanr(p_resid, act_resid)

        print(f"\n  RESULTS ({cell}, n={len(common):,}):")
        print(f"    Raw r                = {r_raw:.3f}")
        print(f"    GC-stratified r      = {r_gc:.3f}  (Table 5 main result)")
        print(f"    GC-corrected residual = {r_corr:.3f}  p={p_corr:.2e}")

        # Random model control
        print(f"  Computing random model (null control) ...", flush=True)
        p_rand = get_p_scores(model_rand, seqs, common, args.device)
        r_rand_gc = gc_stratified_r(p_rand, activity, gc)
        print(f"    Random model GC-strat r = {r_rand_gc:.3f}  (expect ≈ 0)")

    print(f"\n{'='*50}")
    print("Expected results (Table 5 in paper):")
    print("  HepG2: raw r=0.176  GC-stratified r=0.096  n=61,463")
    print("  K562:  raw r=0.124  GC-stratified r=0.110  n=2,123")
    print("  Random model K562: GC-strat r=-0.049")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--checkpoint", required=True)
    parser.add_argument("--seq_table",  required=True,
                        help="Table S3 from Agarwal et al. 2025 (Zenodo 10558183)")
    parser.add_argument("--act_table",  required=True,
                        help="Table S6 from Agarwal et al. 2025")
    parser.add_argument("--cell_type",  default="HepG2",
                        choices=["HepG2", "K562", "all"])
    parser.add_argument("--device",     default="cuda")
    main(parser.parse_args())
