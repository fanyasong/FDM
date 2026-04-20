"""
MPRA correlation analysis: FDM L8 surprisal vs lentiMPRA enhancer activity.

Usage:
    python analysis/mpra_analysis.py \
        --checkpoint checkpoints/entropy_pretrain/best.pt \
        --sequences  data/mpra/table_s3.xlsx \
        --activity   data/mpra/table_s6.xlsx \
        --cell       HepG2

Expected output:
    HepG2: raw r=0.176  GC-stratified r=0.096  n=61,463
    K562:  raw r=0.124  GC-stratified r=0.110  n=2,123
    Random model control: r=-0.049

Data:
    Agarwal, V. et al. Nature 639, 411-420 (2025). Zenodo: 10558183
    Table S3: sequences (200 bp MPRA elements)
    Table S6: log2(RNA/DNA) activity per cell type
"""

import argparse
import sys
from pathlib import Path

import numpy as np
import openpyxl
import torch
from scipy.stats import spearmanr

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from fdm.model import FDM_EntropyLM, load_model
from fdm.utils import cross_layer_normalise

DNA_VOCAB = {'A':0,'T':1,'G':2,'C':3,'N':4,'a':0,'t':1,'g':2,'c':3,'n':4}
SEQ_LEN   = 512   # FDM trained on 512 bp; 200 bp MPRA sequences are N-padded
BATCH     = 128
LAYER     = 7     # L8 (0-indexed)
N_GC_BINS = 5     # GC quintiles for stratification


def load_sequences(seq_xlsx, cell):
    """Load 200 bp MPRA sequences from Agarwal et al. Table S3."""
    sheet_map = {'HepG2': 'HepG2 large-scale', 'K562': 'K562 large-scale'}
    sheet = sheet_map.get(cell, cell)
    wb = openpyxl.load_workbook(str(seq_xlsx), read_only=True)
    ws = wb[sheet]
    seqs = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        name = str(row[0])
        if '_Reversed' in name:
            continue
        seq = None
        for val in row:
            if isinstance(val, str) and len(val) >= 200:
                if set(val.upper()).issubset({'A','T','G','C','N'}):
                    # Remove 15 nt lentiMPRA adaptor if present
                    seq = val[15:215] if len(val) >= 230 else val[:200]
                    break
        if seq:
            seqs[name] = seq
    wb.close()
    return seqs


def load_activity(act_xlsx, cell):
    """Load log2(RNA/DNA) activity from Agarwal et al. Table S6."""
    sheet_map = {
        'HepG2': 'Fold assignment, HepG2',
        'K562':  'Fold assignment, K562',
        'WTC11': 'Fold assignment, WTC11',
    }
    sheet = sheet_map.get(cell, cell)
    wb = openpyxl.load_workbook(str(act_xlsx), read_only=True)
    ws = wb[sheet]
    activity = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[1] is not None and row[2] is not None:
            activity[str(row[1])] = float(row[2])
    wb.close()
    return activity


def get_p_scores(model, seqs_dict, ids, device):
    """Compute L8 normalised surprisal for all sequences."""
    p_out = []
    with torch.no_grad():
        for i in range(0, len(ids), BATCH):
            batch = ids[i:i+BATCH]
            xs = []
            for sid in batch:
                seq = seqs_dict[sid]
                # N-pad to SEQ_LEN
                seq = (seq + 'N' * (SEQ_LEN - len(seq)))[:SEQ_LEN]
                xs.append([DNA_VOCAB.get(b, 4) for b in seq])
            x = torch.tensor(xs, dtype=torch.long).to(device)
            p_vals = model.get_surprisal(x)   # dict layer -> (B, T)
            # Mean over sequence positions, then collect all 8 layers
            layer_means = np.stack(
                [p_vals[li].mean(dim=1).numpy() for li in range(8)],
                axis=1
            )  # (B, 8)
            p_out.extend(layer_means.tolist())
    p_matrix = np.array(p_out)               # (N, 8)
    p_norm = cross_layer_normalise(p_matrix)
    return p_norm[:, LAYER]                  # L8 normalised


def gc_stratified_r(p_scores, activity, gc, n_bins=N_GC_BINS):
    """
    Compute Spearman r within each GC quintile and average.
    This eliminates GC content as a confound.
    """
    gc_bins = np.percentile(gc, np.linspace(0, 100, n_bins + 1))
    rs = []
    for j in range(n_bins):
        mask = (gc >= gc_bins[j]) & (gc < gc_bins[j+1])
        if mask.sum() < 50:
            continue
        r, _ = spearmanr(p_scores[mask], activity[mask])
        rs.append(r)
    return np.mean(rs), rs


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--checkpoint', required=True)
    parser.add_argument('--sequences',  required=True)
    parser.add_argument('--activity',   required=True)
    parser.add_argument('--cell',       default='HepG2', choices=['HepG2', 'K562', 'WTC11'])
    parser.add_argument('--device',     default='cuda' if torch.cuda.is_available() else 'cpu')
    parser.add_argument('--random_control', action='store_true',
                        help='Also run random-model control')
    args = parser.parse_args()

    # Load data
    print(f"Loading {args.cell} sequences...", flush=True)
    seqs = load_sequences(args.sequences, args.cell)
    activity_map = load_activity(args.activity, args.cell)
    common = [k for k in seqs if k in activity_map]
    print(f"  Sequences with activity scores: {len(common):,}")

    activity = np.array([activity_map[k] for k in common])
    gc = np.array([
        (seqs[k].upper().count('G') + seqs[k].upper().count('C')) / 200.0
        for k in common
    ])

    # Trained model
    print("Loading model...", flush=True)
    model = load_model(args.checkpoint, args.device)

    print("Computing surprisal scores...", flush=True)
    p_scores = get_p_scores(model, seqs, common, args.device)

    r_raw, p_raw = spearmanr(p_scores, activity)
    r_gc, rs_by_bin = gc_stratified_r(p_scores, activity, gc)

    print(f"\n{'='*55}")
    print(f"Results: FDM L8 surprisal vs MPRA activity ({args.cell})")
    print(f"{'='*55}")
    print(f"  Raw Spearman r     = {r_raw:.3f}  (p={p_raw:.2e})")
    print(f"  GC-stratified r    = {r_gc:.3f}  (mean across {N_GC_BINS} GC quintiles)")
    print(f"  n                  = {len(common):,}")
    print(f"\n  Per-quintile r: {[f'{r:.3f}' for r in rs_by_bin]}")

    # GC-corrected residual
    gc_coef = np.polyfit(gc, p_scores, 1)
    p_resid = p_scores - (gc_coef[0] * gc + gc_coef[1])
    gc_coef2 = np.polyfit(gc, activity, 1)
    a_resid = activity - (gc_coef2[0] * gc + gc_coef2[1])
    r_corr, p_corr = spearmanr(p_resid, a_resid)
    print(f"  GC-corrected resid r = {r_corr:.3f}  (p={p_corr:.2e})")

    # Random model control
    if args.random_control:
        print("\nRunning random model control...", flush=True)
        model_rand = FDM_EntropyLM().to(args.device)
        model_rand.eval()
        p_rand = get_p_scores(model_rand, seqs, common, args.device)
        r_rand, _ = spearmanr(p_rand, activity)
        r_rand_gc, _ = gc_stratified_r(p_rand, activity, gc)
        print(f"  Random model: raw r={r_rand:.3f}  GC-strat r={r_rand_gc:.3f}")

    print(f"\nExpected ({args.cell}):")
    expected = {'HepG2': 'raw r=0.176  GC-strat r=0.096  n=61,463',
                'K562':  'raw r=0.124  GC-strat r=0.110  n=2,123'}
    print(f"  {expected.get(args.cell, 'see paper')}")


if __name__ == '__main__':
    main()
